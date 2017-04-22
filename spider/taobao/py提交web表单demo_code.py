#-*- coding: UTF-8 -*-   
import urllib2,urllib  
import time
import xlrd
import xlwt
import requests
import cookielib
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

baobei_leimu = "https://sell.xiangqing.taobao.com/sell/start.html?itemId=547105660989&clientType=0&actionStatus=0&frameId=shenbi_1&"
login_url = "https://login.taobao.com/member/login.jhtml"
upload_url = "https://upload.taobao.com/auction/publish/publish.htm"

#base header for use
header = [
  {
    "name": "Cookie",
    "value": "gm_item_upload_start_time=1490261852625; _umdata=70CF403AFFD707DFED08E6D0D10782399C20737F7A4BC46D5700275F728DC4B0AD05CBF6A1A94BFBCD43AD3E795C914CC2C619865BC550D5090C7EFF78F87966; gm_item_upload_take_time=2417817; miid=258847553226394565; thw=cn; UM_distinctid=15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578; _po=50014912_20.0_7_0.0_%E4%B8%8A%E6%B5%B7_%E4%B8%8A%E6%B5%B7_1_%E6%AC%BE%E5%88%B0%E5%8F%91%E8%B4%A7_10000_0_0_5_0_0_0.0_0.0_0.0_true_8940774860_null_0.0_20.0_null_null_null_135554976; x=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0; whl=-1%260%260%261490249632922; v=0; _tb_token_=fb3e5e7777336; uc3=sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWFmrrkZxc%2FsmA%3D&lg2=URm48syIIVrSKA%3D%3D; existShop=MTQ5MDI2MTk5OA%3D%3D; uss=BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D; lgc=qhn614; tracknick=qhn614; cookie2=41ad5598d9dd68684627ac5c611b83ac; sg=464; mt=np=&ci=21_1&cyk=-1_-1; cookie1=AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D; unb=135554976; skt=5540152fe95e625a; t=869b559af8caff3016c5e1c926316a66; _cc_=UIHiLt3xSw%3D%3D; tg=0; _l_g_=Ug%3D%3D; _nk_=qhn614; cookie17=UoLZWZQISEI7; cna=DpmyEIgHin4CAbaKZiKZKYni; uc1=cookie16=Vq8l%2BKCLySLZMFWHxqs8fwqnEw%3D%3D&cookie21=UtASsssmfaCOMId3WwGQmg%3D%3D&cookie15=UtASsssmOIJ0bQ%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKK5jXmvPg%3D%3D&tag=3&lng=zh_CN; l=AiYmhrZ6tlW-rXnt4oWaavTD9paprWrZ; isg=AmZmzec7AiIMO9Zs-euiLoIYt9xWwaoBS3N831AO-Al80wDtt9f6EUzhXXgl"
  },
  {
    "name": "Origin",
    "value": "https://upload.taobao.com"
  },
  {
    "name": "Accept-Encoding",
    "value": "gzip, deflate, br"
  },
  {
    "name": "Host",
    "value": "upload.taobao.com"
  },
  {
    "name": "Accept-Language",
    "value": "zh-CN,zh;q=0.8"
  },
  {
    "name": "Upgrade-Insecure-Requests",
    "value": "1"
  },
  {
    "name": "User-Agent",
    "value": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
  },
  {
    "name": "Content-Type",
    "value": "multipart/form-data; boundary=----WebKitFormBoundaryOtK5kELGehgIvjBB"
  },
  {
    "name": "Accept",
    "value": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8"
  },
  {
    "name": "Cache-Control",
    "value": "max-age=0"
  },
  {
    "name": "Referer",
    "value": "https://upload.taobao.com/auction/publish/publish.htm"
  },
  {
    "name": "Connection",
    "value": "keep-alive"
  },
  {
    "name": "Content-Length",
    "value": "14090"
  }
]

headers = [
#  {
#    "name": "Cookie",
#    "value": 
#  },
  {
    "name": "Origin",
    "value": "https://upload.taobao.com"
  },
  {
    "name": "Accept-Encoding",
    "value": "gzip, deflate, br"
  },
  {
    "name": "Host",
    "value": "upload.taobao.com"
  },
  {
    "name": "Accept-Language",
    "value": "zh-CN,zh;q=0.8"
  },
  {
    "name": "Upgrade-Insecure-Requests",
    "value": "1"
  },
  {
    "name": "User-Agent",
    "value": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
  },
  {
    "name": "Content-Type",
    "value": "multipart/form-data; boundary=----WebKitFormBoundaryOtK5kELGehgIvjBB"
  },
  {
    "name": "Accept",
    "value": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8"
  },
  {
    "name": "Cache-Control",
    "value": "max-age=0"
  },
  {
    "name": "Referer",
    "value": "https://upload.taobao.com/auction/publish/publish.htm"
  },
  {
    "name": "Connection",
    "value": "keep-alive"
  },
  {
    "name": "Content-Length",
    "value": "14090"
  }
]


cookies = [
  {
    "name": "gm_item_upload_start_time",
    "value": "1490261852625",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "_umdata",
    "value": "70CF403AFFD707DFED08E6D0D10782399C20737F7A4BC46D5700275F728DC4B0AD05CBF6A1A94BFBCD43AD3E795C914CC2C619865BC550D5090C7EFF78F87966",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "gm_item_upload_take_time",
    "value": "2417817",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "miid",
    "value": "258847553226394565",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "thw",
    "value": "cn",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "UM_distinctid",
    "value": "15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "_po",
    "value": "50014912_20.0_7_0.0_%E4%B8%8A%E6%B5%B7_%E4%B8%8A%E6%B5%B7_1_%E6%AC%BE%E5%88%B0%E5%8F%91%E8%B4%A7_10000_0_0_5_0_0_0.0_0.0_0.0_true_8940774860_null_0.0_20.0_null_null_null_135554976",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "x",
    "value": "e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "whl",
    "value": "-1%260%260%261490249632922",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "v",
    "value": "0",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "_tb_token_",
    "value": "fb3e5e7777336",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "uc3",
    "value": "sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWFmrrkZxc%2FsmA%3D&lg2=URm48syIIVrSKA%3D%3D",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "existShop",
    "value": "MTQ5MDI2MTk5OA%3D%3D",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "uss",
    "value": "BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "lgc",
    "value": "qhn614",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "tracknick",
    "value": "qhn614",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "cookie2",
    "value": "41ad5598d9dd68684627ac5c611b83ac",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "sg",
    "value": "464",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "mt",
    "value": "np=&ci=21_1&cyk=-1_-1",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "cookie1",
    "value": "AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "unb",
    "value": "135554976",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "skt",
    "value": "5540152fe95e625a",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "t",
    "value": "869b559af8caff3016c5e1c926316a66",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "_cc_",
    "value": "UIHiLt3xSw%3D%3D",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "tg",
    "value": "0",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "_l_g_",
    "value": "Ug%3D%3D",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "_nk_",
    "value": "qhn614",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "cookie17",
    "value": "UoLZWZQISEI7",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "cna",
    "value": "DpmyEIgHin4CAbaKZiKZKYni",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "uc1",
    "value": "cookie16=Vq8l%2BKCLySLZMFWHxqs8fwqnEw%3D%3D&cookie21=UtASsssmfaCOMId3WwGQmg%3D%3D&cookie15=UtASsssmOIJ0bQ%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKK5jXmvPg%3D%3D&tag=3&lng=zh_CN",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "l",
    "value": "AiYmhrZ6tlW-rXnt4oWaavTD9paprWrZ",
    "expires": None,
    "httpOnly": False,
    "secure": False
  },
  {
    "name": "isg",
    "value": "AmZmzec7AiIMO9Zs-euiLoIYt9xWwaoBS3N831AO-Al80wDtt9f6EUzhXXgl",
    "expires": None,
    "httpOnly": False,
    "secure": False
  }
]

post_data = [{
    "name" :"_tb_token_"
    ,
"value" :"fb3e5e7777336"

}
,{
    "name" :"action"
    ,
"value" :"upload/uploadAction"

}
,{
    "name" :"isMImageUser"
    ,
"value" :"True"

}
,{
    "name" :"sellerActionBeginTime"
    ,
"value" :    ""

}
,{
    "name" :"publishPageCostTime"
    ,
"value" :"1490263131301"

}
,{
    "name" :"browserTime"
    ,
"value" :    ""

}
,{
    "name" :"J_BySearch"
    ,
"value" :    ""

}
,{
    "name" :"J_SearchKey"
    ,
"value" :    ""

}
,{
    "name" :"J_ByLateCategory"
    ,
"value" :"1"

}
,{
    "name" :"J_ByCategorySystem"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.ca"
    ,
"value" :"50014912"

}
,{
    "name" :"_fma.pu._0.i"
    ,
"value" :"YTVkN2I4ZjYyYjU5OGVjYWUxYTJlZTNhZDA2MTE4NDAgYjkwODVjMDNkYTI4YWI5NjQyNjg1NmNiMzBlMGMyOTQgMTQ5MDI2MzExMzc3Mw=="

}
,{
    "name" :"_fma.pu._0.auc"
    ,
"value" :"547257315531"

}
,{
    "name" :"itemNumId"
    ,
"value" :"547257315531"

}
,{
    "name" :"_fma.pu._0.cat"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.x"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.ed"
    ,
"value" :"False"

}
,{
    "name" :"isEdit"
    ,
"value" :"False"

}
,{
    "name" :"noticeParam"
    ,
"value" :    ""

}
,{
    "name" :"oldCat"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.fi"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.au"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.fr"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.sho"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.isn"
    ,
"value" :"False"

}
,{
    "name" :"_fma.pu._0.isol"
    ,
"value" :"False"

}
,{
    "name" :"_fma.pu._0.isg"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.ish"
    ,
"value" :"False"

}
,{
    "name" :"catName"
    ,
"value" :"½Å±¾/¹¤¾ß"

}
,{
    "name" :"selPPay"
    ,
"value" :    ""

}
,{
    "name" :"certType"
    ,
"value" :    ""

}
,{
    "name" :"isZFBCertificatedUser"
    ,
"value" :    ""

}
,{
    "name" :"mainCategoryId"
    ,
"value" :"50014811"

}
,{
    "name" :"_fma.pu._0.sp"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.isb"
    ,
"value" :"False"

}
,{
    "name" :"_fma.pu._0.a"
    ,
"value" :"b"

}
,{
    "name" :"_fma.pu._0.pa"
    ,
"value" :"1"

}
,{
    "name" :"_fma.pu._0.b"
    ,
"value" :"20"

}
,{
    "name" :"_fma.pu._0.secure"
    ,
"value" :"1"

}
,{
    "name" :"_fma.pu._0.iss"
    ,
"value" :    ""

}
,{
    "name" :"isCodFlag"
    ,
"value" :    ""

}
,{
    "name" :"ccSn"
    ,
"value" :    ""

}
,{
    "name" :"ccIssuerDn"
    ,
"value" :    ""

}
,{
    "name" :"ccSignedData"
    ,
"value" :    ""

}
,{
    "name" :"userInnerShopId"
    ,
"value" :"122254584"

}
,{
    "name" :"userOuterShopId"
    ,
"value" :    ""

}
,{
    "name" :"userInnerShopSiteId"
    ,
"value" :"4"

}
,{
    "name" :"userOuterShopSiteId"
    ,
"value" :    ""

}
,{
    "name" :"wirelessmpChoosed"
    ,
"value" :    ""

}
,{
    "name" :"isSellingLightningConsignment"
    ,
"value" :    ""

}
,{
    "name" :"lightningConsignment2"
    ,
"value" :    ""

}
,{
    "name" :"spuStoreId"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.ta"
    ,
"value" :"False"

}
,{
    "name" :"newProductTagId"
    ,
"value" :"0"

}
,{
    "name" :"isFirst"
    ,
"value" :    ""

}
,{
    "name" :"tfsname"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.rel"
    ,
"value" :    ""

}
,{
    "name" :"isLocalityLifeEtcItem"
    ,
"value" :    ""

}
,{
    "name" :"isCustoemdSkuForbidDpcAuction"
    ,
"value" :"False"

}
,{
    "name" :"customCatId"
    ,
"value" :    ""

}
,{
    "name" :"customCatName"
    ,
"value" :    ""

}
,{
    "name" :"distributionPageString"
    ,
"value" :    ""

}
,{
    "name" :"isReselectCat"
    ,
"value" :    ""

}
,{
    "name" :"isNeedCheckPictureCat"
    ,
"value" :"False"

}
,{
    "name" :"paymentSeting"
    ,
"value" :    ""

}
,{
    "name" :"forceUpdateWhenFakeCredit"
    ,
"value" :    ""

}
,{
    "name" :"wapDescEnable"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.stu"
    ,
"value" :"5"

}
,{
    "name" :"auctionTypeInfoPrint"
    ,
"value" :"editaVlue=false;canPublishNew=true;isOnlySecond=false;isRestAuction=;isBookStartLimit=false;isAcousticsLimit=false;isPrepayLimitCatAndSeller=false;is3GCatCanNewForSeller=;canPublishAlcoholNew=;noAgreement=;noPrepay=;shopSeller=true;isSellerShopeEnable=true;isShopRelease=;"

}
,{
    "name" :"_fma.pu._0.auctio"
    ,
"value" :"1048330278"

}
,{
    "name" :"customizedProp"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.ti"
    ,
"value" :"²âÊÔÓÃÀý"

}
,{
    "name" :"_fma.pu._0.pay"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.payv"
    ,
"value" :"1"

}
,{
    "name" :"_fma.pu._0.m"
    ,
"value" :"20"

}
,{
    "name" :"_fma.pu._0.q"
    ,
"value" :"10000"

}
,{
    "name" :"_fma.pu._0.q"
    ,
"value" :"10000"

}
,{
    "name" :"gs_country"
    ,
"value" :    ""

}
,{
    "name" :"gs_type"
    ,
"value" :"2"

}
,{
    "name" :"_fma.pu._0.ou"
    ,
"value" :    ""

}
,{
    "name" :"videoAsPicThum"
    ,
"value" :    ""

}
,{
    "name" :"videoAsPicId"
    ,
"value" :    ""

}
,{
    "name" :"videoAsPicDuration"
    ,
"value" :    ""

}
,{
    "name" :"pisAsVideoStatus"
    ,
"value" :    ""

}
,{
    "name" :"image_pos"
    ,
"value" :"1"

}
,{
    "name" :"picUrl1"
    ,
"value" :"https://img.alicdn.com/imgextra/i2/135554976/TB2l66KkohnpuFjSZFpXXcpuXXa_!!135554976.jpg"

}
,{
    "name" :"image_pos"
    ,
"value" :"2"

}
,{
    "name" :"picUrl2"
    ,
"value" :    ""

}
,{
    "name" :"image_pos"
    ,
"value" :3

}
,{
    "name" :"picUrl3"
    ,
"value" :    ""

}
,{
    "name" :"image_pos"
    ,
"value" :4

}
,{
    "name" :"picUrl4"
    ,
"value" :    ""

}
,{
    "name" :"image_pos"
    ,
"value" :5

}
,{
    "name" :"picUrl5"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.is"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.it"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.v"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.sw"
    ,
"value" :    ""

}
,{
    "name" :"handleAuctionModule"
    ,
"value" :"False"

}
,{
    "name" :"modularizedArea"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.pc"
    ,
"value" :"0"

}
,{
    "name" :"_fma.pu._0.d"
    ,
"value" :"²âÊÔÓÃÀý"

}
,{
    "name" :"state"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.w"
    ,
"value" :"0"

}
,{
    "name" :"_fma.pu._0.wi"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.pc"
    ,
"value" :"0"

}
,{
    "name" :"_fma.pu._0.w"
    ,
"value" :"0"

}
,{
    "name" :"pcTspeditorCid"
    ,
"value" :    ""

}
,{
    "name" :"itemNumId"
    ,
"value" :"547257315531"

}
,{
    "name" :"_fma.pu._0.sh"
    ,
"value" :    ""

}
,{
    "name" :"shenbiPcDescHashCode"
    ,
"value" :    ""

}
,{
    "name" :"pcDescbackupEnable"
    ,
"value" :"False"

}
,{
    "name" :"wlDescbackupEnable"
    ,
"value" :"False"

}
,{
    "name" :"isShowShenbiEditor"
    ,
"value" :"True"

}
,{
    "name" :"TspWlsDesc"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.aut"
    ,
"value" :"False"

}
,{
    "name" :"_fma.pu._0.auto"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.autoc"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.po"
    ,
"value" :"8940774860"

}
,{
    "name" :"_fma.pu._0.ha"
    ,
"value" :"0"

}
,{
    "name" :"_fma.pu._0.hav"
    ,
"value" :"0"

}
,{
    "name" :"new_prepay_tag"
    ,
"value" :"0"

}
,{
    "name" :"_fma.pu._0.isop"
    ,
"value" :"False"

}
,{
    "name" :"_fma.pu._0.du"
    ,
"value" :"7"

}
,{
    "name" :"_fma.pu._0.sta"
    ,
"value" :    ""

}
,{
    "name" :"_fma.pu._0.auct"
    ,
"value" :"0"

}
,{
    "name" :"_now"
    ,
"value" :"0"

}
,{
    "name" :"event_submit_do_publish"
    ,
"value" :"anything"

}
,{
    "name" :"stamp"
    ,
"value" :"1490264270553"

}
,{
    "name" :"type"
    ,
"value" :"submit"

}
,{
    "name" :"submitType"
    ,
"value" :"ajax"}
]


class EntryDemo:    
    """Demonstrate Entrys and Event binding"""    
    def __init__(self):
        pass
#        self.url = url
#        self.uid = ''                 #用户名    
#        self.password = ''            # 密码    
#        self.operation = ''           #  操作    
#        self.range = '2'                # 范围    
#        self.the_page = ''            # WEB服务器返回页面    

    # 表单的INPUT 值一定要记得填齐全    
#    def login(self):    
#        postdata = urllib.urlencode(values)                       # 表单值编码
#        req = urllib.Request(self.url_login, postdata)  # 服务器请求    
#        response = urllib.urlopen(req)                 
#        self.the_page = response.read()  

    def uploadwithcookies(self):    
        values = {
          "method": "GET",
          "url": "https://sell.xiangqing.taobao.com/sell/start.html?itemId=547105660989&clientType=0&actionStatus=0&frameId=shenbi_1&",
          "httpVersion": "HTTP/1.1",
          "headers": [
            {
              "name": "Accept-Encoding",
              "value": "gzip, deflate, sdch, br"
            },
            {
              "name": "Host",
              "value": "sell.xiangqing.taobao.com"
            },
            {
              "name": "Accept-Language",
              "value": "zh-CN,zh;q=0.8"
            },
            {
              "name": "Upgrade-Insecure-Requests",
              "value": "1"
            },
            {
              "name": "User-Agent",
              "value": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
            },
            {
              "name": "Accept",
              "value": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8"
            },
            {
              "name": "Referer",
              "value": "https://upload.taobao.com/auction/publish/publish.htm"
            },
            {
              "name": "Cookie",
              "value": "miid=258847553226394565; thw=cn; UM_distinctid=15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578; x=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0; v=0; _tb_token_=fb3e5e7777336; uc3=sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWKWIl6fa4wXBQ%3D&lg2=WqG3DMC9VAQiUQ%3D%3D; existShop=MTQ5MDE2NjYwMw%3D%3D; uss=BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D; lgc=qhn614; tracknick=qhn614; cookie2=41ad5598d9dd68684627ac5c611b83ac; sg=464; mt=np=&ci=21_1&cyk=-1_-1; cookie1=AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D; unb=135554976; skt=2167bce167878caa; t=869b559af8caff3016c5e1c926316a66; _cc_=VFC%2FuZ9ajQ%3D%3D; tg=0; _l_g_=Ug%3D%3D; _nk_=qhn614; cookie17=UoLZWZQISEI7; cna=DpmyEIgHin4CAbaKZiKZKYni; l=AkZGK2btlrVeTVlNwmW6i0BkFjbJ0Ipg; isg=Alxc62u7aL0GrhwKn9k44MwqLXra8QD_tYG2eTZcH8crgf8LXuQyjhRxl16D; uc1=cookie16=U%2BGCWk%2F74Mx5tgzv3dWpnhjPaQ%3D%3D&cookie21=VFC%2FuZ9ajCbF8%2BYBpbBdiw%3D%3D&cookie15=UIHiLt3xD8xYTw%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKEsdspHkg%3D%3D&tag=3&lng=zh_CN"
            },
            {
              "name": "Connection",
              "value": "keep-alive"
            }
          ],
          "queryString": [
            {
              "name": "itemId",
              "value": "547105660989"
            },
            {
              "name": "clientType",
              "value": "0"
            },
            {
              "name": "actionStatus",
              "value": "0"
            },
            {
              "name": "frameId",
              "value": "shenbi_1"
            },
            {
              "name": "",
              "value": ""
            }
          ],
          "cookies": [
            {
              "name": "miid",
              "value": "258847553226394565",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "thw",
              "value": "cn",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "UM_distinctid",
              "value": "15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "x",
              "value": "e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "v",
              "value": "0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_tb_token_",
              "value": "fb3e5e7777336",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uc3",
              "value": "sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWKWIl6fa4wXBQ%3D&lg2=WqG3DMC9VAQiUQ%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "existShop",
              "value": "MTQ5MDE2NjYwMw%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uss",
              "value": "BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "lgc",
              "value": "qhn614",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "tracknick",
              "value": "qhn614",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cookie2",
              "value": "41ad5598d9dd68684627ac5c611b83ac",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "sg",
              "value": "464",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "mt",
              "value": "np=&ci=21_1&cyk=-1_-1",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cookie1",
              "value": "AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "unb",
              "value": "135554976",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "skt",
              "value": "2167bce167878caa",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "t",
              "value": "869b559af8caff3016c5e1c926316a66",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_cc_",
              "value": "VFC%2FuZ9ajQ%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "tg",
              "value": "0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_l_g_",
              "value": "Ug%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_nk_",
              "value": "qhn614",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cookie17",
              "value": "UoLZWZQISEI7",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cna",
              "value": "DpmyEIgHin4CAbaKZiKZKYni",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "l",
              "value": "AkZGK2btlrVeTVlNwmW6i0BkFjbJ0Ipg",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "isg",
              "value": "Alxc62u7aL0GrhwKn9k44MwqLXra8QD_tYG2eTZcH8crgf8LXuQyjhRxl16D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uc1",
              "value": "cookie16=U%2BGCWk%2F74Mx5tgzv3dWpnhjPaQ%3D%3D&cookie21=VFC%2FuZ9ajCbF8%2BYBpbBdiw%3D%3D&cookie15=UIHiLt3xD8xYTw%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKEsdspHkg%3D%3D&tag=3&lng=zh_CN",
              "expires": None,
              "httpOnly": False,
              "secure": False
            }
          ],
          "headersSize": 1642,
          "bodySize": 0
        }
        postdata = urllib.urlencode(values)                       # 表单值编码
        response = urllib.urlopen(baobei_leimu, postdata)                 
        self.the_page = response.read()  
        print (self.the_page)

    def leimu_test(self):
    
        values={
          "method": "POST",
          "httpVersion": "HTTP/1.1",
          "headers": [
            {
              "name": "Cookie",
              "value": "_umdata=70CF403AFFD707DFED08E6D0D10782399C20737F7A4BC46D5700275F728DC4B0AD05CBF6A1A94BFBCD43AD3E795C914CC2C619865BC550D5090C7EFF78F87966; miid=258847553226394565; thw=cn; UM_distinctid=15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578; _po=50014912_20.0_7_0.0_%E4%B8%8A%E6%B5%B7_%E4%B8%8A%E6%B5%B7_1_%E6%AC%BE%E5%88%B0%E5%8F%91%E8%B4%A7_10000_0_0_5_0_0_0.0_0.0_0.0_true_8940774860_null_0.0_20.0_null_null_null_135554976; x=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0; whl=-1%260%260%261490249632922; v=0; _tb_token_=fb3e5e7777336; uc3=sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWFmrrkZxc%2FsmA%3D&lg2=URm48syIIVrSKA%3D%3D; existShop=MTQ5MDI2MTk5OA%3D%3D; uss=BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D; lgc=qhn614; tracknick=qhn614; cookie2=41ad5598d9dd68684627ac5c611b83ac; sg=464; mt=np=&ci=21_1&cyk=-1_-1; cookie1=AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D; unb=135554976; skt=5540152fe95e625a; t=869b559af8caff3016c5e1c926316a66; _cc_=UIHiLt3xSw%3D%3D; tg=0; _l_g_=Ug%3D%3D; _nk_=qhn614; cookie17=UoLZWZQISEI7; cna=DpmyEIgHin4CAbaKZiKZKYni; l=AkFBubaKCXSdhH5csfx1K4Bt0YdbybVj; isg=Am1tOPAmyR8M-a2ZZrY5Y_WNfAlZTKGcHAanyq9yAYRwJo7YdxiNbeW0Zk06; uc1=cookie16=VT5L2FSpNgq6fDudInPRgavC%2BQ%3D%3D&cookie21=Vq8l%2BKCLiv0MyZ1zjQnMQw%3D%3D&cookie15=W5iHLLyFOGW7aA%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKK5i4pLcw%3D%3D&tag=3&lng=zh_CN"
            },
            {
              "name": "Origin",
              "value": "https://upload.taobao.com"
            },
            {
              "name": "Accept-Encoding",
              "value": "gzip, deflate, br"
            },
            {
              "name": "Host",
              "value": "upload.taobao.com"
            },
            {
              "name": "Accept-Language",
              "value": "zh-CN,zh;q=0.8"
            },
            {
              "name": "Upgrade-Insecure-Requests",
              "value": "1"
            },
            {
              "name": "User-Agent",
              "value": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
            },
            {
              "name": "Content-Type",
              "value": "application/x-www-form-urlencoded"
            },
            {
              "name": "Accept",
              "value": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8"
            },
            {
              "name": "Cache-Control",
              "value": "max-age=0"
            },
            {
              "name": "Referer",
              "value": "https://upload.taobao.com/auction/publish/publish.htm"
            },
            {
              "name": "Connection",
              "value": "keep-alive"
            },
            {
              "name": "Content-Length",
              "value": "4180"
            }
          ],
          "queryString": [],
          "cookies": [
            {
              "name": "_umdata",
              "value": "70CF403AFFD707DFED08E6D0D10782399C20737F7A4BC46D5700275F728DC4B0AD05CBF6A1A94BFBCD43AD3E795C914CC2C619865BC550D5090C7EFF78F87966",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "miid",
              "value": "258847553226394565",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "thw",
              "value": "cn",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "UM_distinctid",
              "value": "15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_po",
              "value": "50014912_20.0_7_0.0_%E4%B8%8A%E6%B5%B7_%E4%B8%8A%E6%B5%B7_1_%E6%AC%BE%E5%88%B0%E5%8F%91%E8%B4%A7_10000_0_0_5_0_0_0.0_0.0_0.0_true_8940774860_null_0.0_20.0_null_null_null_135554976",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "x",
              "value": "e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "whl",
              "value": "-1%260%260%261490249632922",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "v",
              "value": "0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_tb_token_",
              "value": "fb3e5e7777336",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uc3",
              "value": "sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWFmrrkZxc%2FsmA%3D&lg2=URm48syIIVrSKA%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "existShop",
              "value": "MTQ5MDI2MTk5OA%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uss",
              "value": "BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "lgc",
              "value": "qhn614",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "tracknick",
              "value": "qhn614",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cookie2",
              "value": "41ad5598d9dd68684627ac5c611b83ac",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "sg",
              "value": "464",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "mt",
              "value": "np=&ci=21_1&cyk=-1_-1",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cookie1",
              "value": "AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "unb",
              "value": "135554976",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "skt",
              "value": "5540152fe95e625a",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "t",
              "value": "869b559af8caff3016c5e1c926316a66",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_cc_",
              "value": "UIHiLt3xSw%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "tg",
              "value": "0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_l_g_",
              "value": "Ug%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_nk_",
              "value": "qhn614",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cookie17",
              "value": "UoLZWZQISEI7",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cna",
              "value": "DpmyEIgHin4CAbaKZiKZKYni",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "l",
              "value": "AkFBubaKCXSdhH5csfx1K4Bt0YdbybVj",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "isg",
              "value": "Am1tOPAmyR8M-a2ZZrY5Y_WNfAlZTKGcHAanyq9yAYRwJo7YdxiNbeW0Zk06",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uc1",
              "value": "cookie16=VT5L2FSpNgq6fDudInPRgavC%2BQ%3D%3D&cookie21=Vq8l%2BKCLiv0MyZ1zjQnMQw%3D%3D&cookie15=W5iHLLyFOGW7aA%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKK5i4pLcw%3D%3D&tag=3&lng=zh_CN",
              "expires": None,
              "httpOnly": False,
              "secure": False
            }
          ],
          "headersSize": 2058,
          "bodySize": 4180,
          "postData": {
            "mimeType": "application/x-www-form-urlencoded",
            "text": "sellerActionBeginTime=&doReSelectCategoryLimit=&J_BySearch=0&J_ByLateCategory=1&J_ByCategorySystem=0&J_SearchKey=&_fma.pu._0.ca=50014912&_fma.pu._0.cat=&action=upload%2FuploadAction&event_submit_do_input_auction_info=1&_fma.pu._0.i=MGYzNjVkMzJlOGRlOTk0N2Y2MzM1MDBjZWYzY2E1MDUgYjkwODVjMDNkYTI4YWI5NjQyNjg1NmNiMzBlMGMyOTQgMTQ5MDI2MjAwNjE5OQ%3D%3D&_fma.pu._0.auc=547175536778&_fma.pu._0.x=&_fma.pu._0.a=b&_fma.pu._0.ti=&_fma.pu._0.pi=&_fma.pu._0.wi=&_fma.e._0.t=&_fma.e._0.p=&_fma.e._0.i=&_fma.e._0.s=&_fma.e._0.it=&_fma.e._0.d=&_fma.e._0.o=&_fma.e._0.is=publish&_fma.e._0.ise=&_fma.e._0.b=&_fma.e._0.a=&_fma.e._0.se=&_fma.e._0.iso=&_fma.e._0.m=&_fma.e._0.pu=&_fma.e._0.in=&_fma.e._0.am=&_fma.e._0.cu=&isEdit=false&isHaveImageExtra=&_fma.pu._0.d=&_fma.pu._0.pc=&_fma.pu._0.w=&_fma.pu._0.sh=&_fma.pu._0.she=&_fma.pu._0.q=10000&_fma.pu._0.du=7&_fma.pu._0.stu=5&_fma.pu._0.m=&_fma.pu._0.bu=5&%24group.incrementNum.key=&_fma.pu._0.pro=%C9%CF%BA%A3&_fma.pu._0.ci=%C9%CF%BA%A3&_fma.pu._0.shi=1&_fma.pu._0.sec=0.0&_fma.pu._0.secu=0.0&_fma.pu._0.secur=0.0&_fma.pu._0.pa=%BF%EE%B5%BD%B7%A2%BB%F5&_fma.pu._0.ha=0&_fma.pu._0.hav=0&_fma.pu._0.secure=0&_fma.pu._0.auct=0&_fma.pu._0.sta=&_fma.pu._0.ed=false&_fma.pu._0.fr=&_fma.pu._0.sho=&_fma.pu._0.aut=false&_fma.pu._0.b=20.0&oldDesc=&oldStory=&productRootCat=&productId=&oldCat=&oldSpuId=&oldCategoryProperty=&_fma.pu._0.isa=false&_fma.pu._0.po=0&_fma.pu._0.an=false&_fma.pu._0.ou=&_fma.pu._0.aucti=0.0&_fma.pu._0.buy=&_fma.pu._0.ec=&_fma.pu._0.eca=&_fma.pu._0.n=&_fma.pu._0.isn=false&_fma.pu._0.isol=false&isZFBCertificatedUser=&ccSn=&ccIssuerDn=&ccSignedData=&_fma.pu._0.st=&_fma.pu._0.sk=&_fma.pu._0.gtr=&_fma.pu._0.g=&%24group.protocol.key=&_fma.pu._0.su=1&_fma.pu._0.auctio=&_fma.pu._0.auction=&userInnerShopId=&userOuterShopId=&userInnerShopSiteId=&userOuterShopSiteId=&itemBarCode=&_fma.pu._0.isap=false&_fma.pu._0.co=&wirelessmpChoosed=&wirelessOption=&_fma.pu._0.se=&_fma.pu._0.sel=false&_fma.pu._0.qu=&_fma.pu._0.l=1&isSellingLightningConsignment=&lightningConsignment2=&isCustoemdSkuForbidDpcAuction=&tangramOriginalDesc=&_fma.pu._0.auctionm=0&%24group.auctionDepositMode.key=%24group.auctionDepositMode.value&%24group.auctionDepositModeDesc.key=%24group.auctionDepositModeDesc.value&%24group.auctionFixedDepositValue.key=%24group.auctionFixedDepositValue.value&_fma.pu._0.rel=&distributionPageString=&distributionAssId=&_fma.pu._0.etc=&_fma.pu._0.etce=&_fma.pu._0.exp=1&_fma.pu._0.etcf=&_fma.pu._0.etcex=&_fma.pu._0.ve=0&_fma.pu._0.ov=&etc_network_id=&etc_obs=&etc_merchant_id=&etc_merchant_nick=&etc_has_pos=&etc_eticket=&etc_pool_id=&etc_pool_size=&wapDescEnable=&_fma.pu._0.pay=&_fma.pu._0.payv=1&hasExtendForTicketItem=&_fma.pu._0.fa=&_fma.pu._0.far=&_fma.pu._0.farm=&_fma.pu._0.farmr=&isReselectCat=&_fma.pu._0.etcde=&_fma.pu._0.etcco=&_fma.pu._0.etcc=&_fma.pu._0.etcd=&customCatId=&customCatName=&videoAsPicThum=&videoAsPicId=&videoAsPicDuration=&pisAsVideoStatus=&subTitle=&frontPrice=&voucherPrice=&referencePrice=&_fma.pu._0.iso=&isDeductible=&installment=&noticeParam=&itemNumId=547175536778&itemId=&title=&picUrl1=&picUrl2=&picUrl3=&picUrl4=&picUrl5=&picUrl6=&old_image_1=&isvId=&itemVideoId=&videoId=&description=&shopCategoriesIdList=&prov=&city=&postageid=&haveInvoice=&deliveryEtc=&deliveryLogis=&verificationPay=&refund_check=&overduePay=&autoRefundCheck=&autoRefund=&etc_network_id=&etc_obs=&etc_merchant_id=&etc_merchant_nick=&etc_coupon=&isOptionPromoted=&subStockAtBuy=&auctionStatus=&_now=&_date=&_hour=&_minute=&auctionPoint=&promotedStatus=&guide=&orderLimit=&t_name1=&t_name2=&t_name3=&t_price1=&t_price2=&t_price3=&_fma.pu._0.ex=&is_global_stock=&gs_country=&gs_type=&shenbiPcDescHashCode=&shenbiWlDescHashCode=&_fma.pu._0.cou=&_fma.pu._0.cour=&_fma.pu._0.cours=&_fma.pu._0.issu=false&%24%7Bgroup.foodTitle.key%7D=%24%7Bgroup.foodTitle.value%7D&%24%7Bgroup.customsNumber.key%7D=%24%7Bgroup.customsNumber.value%7D&%24%7Bgroup.edeclarationPicUrl.key%7D=%24%7Bgroup.edeclarationPicUrl.value%7D&%24%7Bgroup.healthyNumber.key%7D=%24%7Bgroup.healthyNumber.value%7D&%24%7Bgroup.healthyPic.key%7D=%24%7Bgroup.healthyPic.value%7D&draftId=&cfProductId=&fenxiaoProduct=%24params.fenxiaoProduct&_tb_token_=fb3e5e7777336",
            "params": [
              {
                "name": "sellerActionBeginTime",
                "value": ""
              },
              {
                "name": "doReSelectCategoryLimit",
                "value": ""
              },
              {
                "name": "J_BySearch",
                "value": "0"
              },
              {
                "name": "J_ByLateCategory",
                "value": "1"
              },
              {
                "name": "J_ByCategorySystem",
                "value": "0"
              },
              {
                "name": "J_SearchKey",
                "value": ""
              },
              {
                "name": "_fma.pu._0.ca",
                "value": "50014912"
              },
              {
                "name": "_fma.pu._0.cat",
                "value": ""
              },
              {
                "name": "action",
                "value": "upload%2FuploadAction"
              },
              {
                "name": "event_submit_do_input_auction_info",
                "value": "1"
              },
              {
                "name": "_fma.pu._0.i",
                "value": "MGYzNjVkMzJlOGRlOTk0N2Y2MzM1MDBjZWYzY2E1MDUgYjkwODVjMDNkYTI4YWI5NjQyNjg1NmNiMzBlMGMyOTQgMTQ5MDI2MjAwNjE5OQ%3D%3D"
              },
              {
                "name": "_fma.pu._0.auc",
                "value": "547175536778"
              },
              {
                "name": "_fma.pu._0.x",
                "value": ""
              },
              {
                "name": "_fma.pu._0.a",
                "value": "b"
              },
              {
                "name": "_fma.pu._0.ti",
                "value": ""
              },
              {
                "name": "_fma.pu._0.pi",
                "value": ""
              },
              {
                "name": "_fma.pu._0.wi",
                "value": ""
              },
              {
                "name": "_fma.e._0.t",
                "value": ""
              },
              {
                "name": "_fma.e._0.p",
                "value": ""
              },
              {
                "name": "_fma.e._0.i",
                "value": ""
              },
              {
                "name": "_fma.e._0.s",
                "value": ""
              },
              {
                "name": "_fma.e._0.it",
                "value": ""
              },
              {
                "name": "_fma.e._0.d",
                "value": ""
              },
              {
                "name": "_fma.e._0.o",
                "value": ""
              },
              {
                "name": "_fma.e._0.is",
                "value": "publish"
              },
              {
                "name": "_fma.e._0.ise",
                "value": ""
              },
              {
                "name": "_fma.e._0.b",
                "value": ""
              },
              {
                "name": "_fma.e._0.a",
                "value": ""
              },
              {
                "name": "_fma.e._0.se",
                "value": ""
              },
              {
                "name": "_fma.e._0.iso",
                "value": ""
              },
              {
                "name": "_fma.e._0.m",
                "value": ""
              },
              {
                "name": "_fma.e._0.pu",
                "value": ""
              },
              {
                "name": "_fma.e._0.in",
                "value": ""
              },
              {
                "name": "_fma.e._0.am",
                "value": ""
              },
              {
                "name": "_fma.e._0.cu",
                "value": ""
              },
              {
                "name": "isEdit",
                "value": "false"
              },
              {
                "name": "isHaveImageExtra",
                "value": ""
              },
              {
                "name": "_fma.pu._0.d",
                "value": ""
              },
              {
                "name": "_fma.pu._0.pc",
                "value": ""
              },
              {
                "name": "_fma.pu._0.w",
                "value": ""
              },
              {
                "name": "_fma.pu._0.sh",
                "value": ""
              },
              {
                "name": "_fma.pu._0.she",
                "value": ""
              },
              {
                "name": "_fma.pu._0.q",
                "value": "10000"
              },
              {
                "name": "_fma.pu._0.du",
                "value": "7"
              },
              {
                "name": "_fma.pu._0.stu",
                "value": "5"
              },
              {
                "name": "_fma.pu._0.m",
                "value": ""
              },
              {
                "name": "_fma.pu._0.bu",
                "value": "5"
              },
              {
                "name": "%24group.incrementNum.key",
                "value": ""
              },
              {
                "name": "_fma.pu._0.pro",
                "value": "%C9%CF%BA%A3"
              },
              {
                "name": "_fma.pu._0.ci",
                "value": "%C9%CF%BA%A3"
              },
              {
                "name": "_fma.pu._0.shi",
                "value": "1"
              },
              {
                "name": "_fma.pu._0.sec",
                "value": "0.0"
              },
              {
                "name": "_fma.pu._0.secu",
                "value": "0.0"
              },
              {
                "name": "_fma.pu._0.secur",
                "value": "0.0"
              },
              {
                "name": "_fma.pu._0.pa",
                "value": "%BF%EE%B5%BD%B7%A2%BB%F5"
              },
              {
                "name": "_fma.pu._0.ha",
                "value": "0"
              },
              {
                "name": "_fma.pu._0.hav",
                "value": "0"
              },
              {
                "name": "_fma.pu._0.secure",
                "value": "0"
              },
              {
                "name": "_fma.pu._0.auct",
                "value": "0"
              },
              {
                "name": "_fma.pu._0.sta",
                "value": ""
              },
              {
                "name": "_fma.pu._0.ed",
                "value": "false"
              },
              {
                "name": "_fma.pu._0.fr",
                "value": ""
              },
              {
                "name": "_fma.pu._0.sho",
                "value": ""
              },
              {
                "name": "_fma.pu._0.aut",
                "value": "false"
              },
              {
                "name": "_fma.pu._0.b",
                "value": "20.0"
              },
              {
                "name": "oldDesc",
                "value": ""
              },
              {
                "name": "oldStory",
                "value": ""
              },
              {
                "name": "productRootCat",
                "value": ""
              },
              {
                "name": "productId",
                "value": ""
              },
              {
                "name": "oldCat",
                "value": ""
              },
              {
                "name": "oldSpuId",
                "value": ""
              },
              {
                "name": "oldCategoryProperty",
                "value": ""
              },
              {
                "name": "_fma.pu._0.isa",
                "value": "false"
              },
              {
                "name": "_fma.pu._0.po",
                "value": "0"
              },
              {
                "name": "_fma.pu._0.an",
                "value": "false"
              },
              {
                "name": "_fma.pu._0.ou",
                "value": ""
              },
              {
                "name": "_fma.pu._0.aucti",
                "value": "0.0"
              },
              {
                "name": "_fma.pu._0.buy",
                "value": ""
              },
              {
                "name": "_fma.pu._0.ec",
                "value": ""
              },
              {
                "name": "_fma.pu._0.eca",
                "value": ""
              },
              {
                "name": "_fma.pu._0.n",
                "value": ""
              },
              {
                "name": "_fma.pu._0.isn",
                "value": "false"
              },
              {
                "name": "_fma.pu._0.isol",
                "value": "false"
              },
              {
                "name": "isZFBCertificatedUser",
                "value": ""
              },
              {
                "name": "ccSn",
                "value": ""
              },
              {
                "name": "ccIssuerDn",
                "value": ""
              },
              {
                "name": "ccSignedData",
                "value": ""
              },
              {
                "name": "_fma.pu._0.st",
                "value": ""
              },
              {
                "name": "_fma.pu._0.sk",
                "value": ""
              },
              {
                "name": "_fma.pu._0.gtr",
                "value": ""
              },
              {
                "name": "_fma.pu._0.g",
                "value": ""
              },
              {
                "name": "%24group.protocol.key",
                "value": ""
              },
              {
                "name": "_fma.pu._0.su",
                "value": "1"
              },
              {
                "name": "_fma.pu._0.auctio",
                "value": ""
              },
              {
                "name": "_fma.pu._0.auction",
                "value": ""
              },
              {
                "name": "userInnerShopId",
                "value": ""
              },
              {
                "name": "userOuterShopId",
                "value": ""
              },
              {
                "name": "userInnerShopSiteId",
                "value": ""
              },
              {
                "name": "userOuterShopSiteId",
                "value": ""
              },
              {
                "name": "itemBarCode",
                "value": ""
              },
              {
                "name": "_fma.pu._0.isap",
                "value": "false"
              },
              {
                "name": "_fma.pu._0.co",
                "value": ""
              },
              {
                "name": "wirelessmpChoosed",
                "value": ""
              },
              {
                "name": "wirelessOption",
                "value": ""
              },
              {
                "name": "_fma.pu._0.se",
                "value": ""
              },
              {
                "name": "_fma.pu._0.sel",
                "value": "false"
              },
              {
                "name": "_fma.pu._0.qu",
                "value": ""
              },
              {
                "name": "_fma.pu._0.l",
                "value": "1"
              },
              {
                "name": "isSellingLightningConsignment",
                "value": ""
              },
              {
                "name": "lightningConsignment2",
                "value": ""
              },
              {
                "name": "isCustoemdSkuForbidDpcAuction",
                "value": ""
              },
              {
                "name": "tangramOriginalDesc",
                "value": ""
              },
              {
                "name": "_fma.pu._0.auctionm",
                "value": "0"
              },
              {
                "name": "%24group.auctionDepositMode.key",
                "value": "%24group.auctionDepositMode.value"
              },
              {
                "name": "%24group.auctionDepositModeDesc.key",
                "value": "%24group.auctionDepositModeDesc.value"
              },
              {
                "name": "%24group.auctionFixedDepositValue.key",
                "value": "%24group.auctionFixedDepositValue.value"
              },
              {
                "name": "_fma.pu._0.rel",
                "value": ""
              },
              {
                "name": "distributionPageString",
                "value": ""
              },
              {
                "name": "distributionAssId",
                "value": ""
              },
              {
                "name": "_fma.pu._0.etc",
                "value": ""
              },
              {
                "name": "_fma.pu._0.etce",
                "value": ""
              },
              {
                "name": "_fma.pu._0.exp",
                "value": "1"
              },
              {
                "name": "_fma.pu._0.etcf",
                "value": ""
              },
              {
                "name": "_fma.pu._0.etcex",
                "value": ""
              },
              {
                "name": "_fma.pu._0.ve",
                "value": "0"
              },
              {
                "name": "_fma.pu._0.ov",
                "value": ""
              },
              {
                "name": "etc_network_id",
                "value": ""
              },
              {
                "name": "etc_obs",
                "value": ""
              },
              {
                "name": "etc_merchant_id",
                "value": ""
              },
              {
                "name": "etc_merchant_nick",
                "value": ""
              },
              {
                "name": "etc_has_pos",
                "value": ""
              },
              {
                "name": "etc_eticket",
                "value": ""
              },
              {
                "name": "etc_pool_id",
                "value": ""
              },
              {
                "name": "etc_pool_size",
                "value": ""
              },
              {
                "name": "wapDescEnable",
                "value": ""
              },
              {
                "name": "_fma.pu._0.pay",
                "value": ""
              },
              {
                "name": "_fma.pu._0.payv",
                "value": "1"
              },
              {
                "name": "hasExtendForTicketItem",
                "value": ""
              },
              {
                "name": "_fma.pu._0.fa",
                "value": ""
              },
              {
                "name": "_fma.pu._0.far",
                "value": ""
              },
              {
                "name": "_fma.pu._0.farm",
                "value": ""
              },
              {
                "name": "_fma.pu._0.farmr",
                "value": ""
              },
              {
                "name": "isReselectCat",
                "value": ""
              },
              {
                "name": "_fma.pu._0.etcde",
                "value": ""
              },
              {
                "name": "_fma.pu._0.etcco",
                "value": ""
              },
              {
                "name": "_fma.pu._0.etcc",
                "value": ""
              },
              {
                "name": "_fma.pu._0.etcd",
                "value": ""
              },
              {
                "name": "customCatId",
                "value": ""
              },
              {
                "name": "customCatName",
                "value": ""
              },
              {
                "name": "videoAsPicThum",
                "value": ""
              },
              {
                "name": "videoAsPicId",
                "value": ""
              },
              {
                "name": "videoAsPicDuration",
                "value": ""
              },
              {
                "name": "pisAsVideoStatus",
                "value": ""
              },
              {
                "name": "subTitle",
                "value": ""
              },
              {
                "name": "frontPrice",
                "value": ""
              },
              {
                "name": "voucherPrice",
                "value": ""
              },
              {
                "name": "referencePrice",
                "value": ""
              },
              {
                "name": "_fma.pu._0.iso",
                "value": ""
              },
              {
                "name": "isDeductible",
                "value": ""
              },
              {
                "name": "installment",
                "value": ""
              },
              {
                "name": "noticeParam",
                "value": ""
              },
              {
                "name": "itemNumId",
                "value": "547175536778"
              },
              {
                "name": "itemId",
                "value": ""
              },
              {
                "name": "title",
                "value": ""
              },
              {
                "name": "picUrl1",
                "value": ""
              },
              {
                "name": "picUrl2",
                "value": ""
              },
              {
                "name": "picUrl3",
                "value": ""
              },
              {
                "name": "picUrl4",
                "value": ""
              },
              {
                "name": "picUrl5",
                "value": ""
              },
              {
                "name": "picUrl6",
                "value": ""
              },
              {
                "name": "old_image_1",
                "value": ""
              },
              {
                "name": "isvId",
                "value": ""
              },
              {
                "name": "itemVideoId",
                "value": ""
              },
              {
                "name": "videoId",
                "value": ""
              },
              {
                "name": "description",
                "value": ""
              },
              {
                "name": "shopCategoriesIdList",
                "value": ""
              },
              {
                "name": "prov",
                "value": ""
              },
              {
                "name": "city",
                "value": ""
              },
              {
                "name": "postageid",
                "value": ""
              },
              {
                "name": "haveInvoice",
                "value": ""
              },
              {
                "name": "deliveryEtc",
                "value": ""
              },
              {
                "name": "deliveryLogis",
                "value": ""
              },
              {
                "name": "verificationPay",
                "value": ""
              },
              {
                "name": "refund_check",
                "value": ""
              },
              {
                "name": "overduePay",
                "value": ""
              },
              {
                "name": "autoRefundCheck",
                "value": ""
              },
              {
                "name": "autoRefund",
                "value": ""
              },
              {
                "name": "etc_network_id",
                "value": ""
              },
              {
                "name": "etc_obs",
                "value": ""
              },
              {
                "name": "etc_merchant_id",
                "value": ""
              },
              {
                "name": "etc_merchant_nick",
                "value": ""
              },
              {
                "name": "etc_coupon",
                "value": ""
              },
              {
                "name": "isOptionPromoted",
                "value": ""
              },
              {
                "name": "subStockAtBuy",
                "value": ""
              },
              {
                "name": "auctionStatus",
                "value": ""
              },
              {
                "name": "_now",
                "value": ""
              },
              {
                "name": "_date",
                "value": ""
              },
              {
                "name": "_hour",
                "value": ""
              },
              {
                "name": "_minute",
                "value": ""
              },
              {
                "name": "auctionPoint",
                "value": ""
              },
              {
                "name": "promotedStatus",
                "value": ""
              },
              {
                "name": "guide",
                "value": ""
              },
              {
                "name": "orderLimit",
                "value": ""
              },
              {
                "name": "t_name1",
                "value": ""
              },
              {
                "name": "t_name2",
                "value": ""
              },
              {
                "name": "t_name3",
                "value": ""
              },
              {
                "name": "t_price1",
                "value": ""
              },
              {
                "name": "t_price2",
                "value": ""
              },
              {
                "name": "t_price3",
                "value": ""
              },
              {
                "name": "_fma.pu._0.ex",
                "value": ""
              },
              {
                "name": "is_global_stock",
                "value": ""
              },
              {
                "name": "gs_country",
                "value": ""
              },
              {
                "name": "gs_type",
                "value": ""
              },
              {
                "name": "shenbiPcDescHashCode",
                "value": ""
              },
              {
                "name": "shenbiWlDescHashCode",
                "value": ""
              },
              {
                "name": "_fma.pu._0.cou",
                "value": ""
              },
              {
                "name": "_fma.pu._0.cour",
                "value": ""
              },
              {
                "name": "_fma.pu._0.cours",
                "value": ""
              },
              {
                "name": "_fma.pu._0.issu",
                "value": "false"
              },
              {
                "name": "%24%7Bgroup.foodTitle.key%7D",
                "value": "%24%7Bgroup.foodTitle.value%7D"
              },
              {
                "name": "%24%7Bgroup.customsNumber.key%7D",
                "value": "%24%7Bgroup.customsNumber.value%7D"
              },
              {
                "name": "%24%7Bgroup.edeclarationPicUrl.key%7D",
                "value": "%24%7Bgroup.edeclarationPicUrl.value%7D"
              },
              {
                "name": "%24%7Bgroup.healthyNumber.key%7D",
                "value": "%24%7Bgroup.healthyNumber.value%7D"
              },
              {
                "name": "%24%7Bgroup.healthyPic.key%7D",
                "value": "%24%7Bgroup.healthyPic.value%7D"
              },
              {
                "name": "draftId",
                "value": ""
              },
              {
                "name": "cfProductId",
                "value": ""
              },
              {
                "name": "fenxiaoProduct",
                "value": "%24params.fenxiaoProduct"
              },
              {
                "name": "_tb_token_",
                "value": "fb3e5e7777336"
              }
            ]
          }
        }    
        
#        POST /auction/publish/publish.htm HTTP/1.1
        values2 = {
          "Host": "upload.taobao.com",
          "Connection": "keep-alive",
          "Content-Length": "4180",
          "Cache-Control": "max-age=0",
          "Origin": "https://upload.taobao.com",
          "Upgrade-Insecure-Requests": "1",
          "User-Agent": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko), Chrome/56.0.2924.87 Safari/537.36",
          "Content-Type": "application/x-www-form-urlencoded",
          "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
          "Referer": "https://upload.taobao.com/auction/publish/publish.htm",
          "Accept-Encoding": "gzip, deflate, br",
          "Accept-Language": "zh-CN,zh;q=0.8",
          "Cookie": "gm_item_upload_start_time=1490261852625; _umdata=70CF403AFFD707DFED08E6D0D10782399C20737F7A4BC46D5700275F728DC4B0AD05CBF6A1A94BFBCD43AD3E795C914CC2C619865BC550D5090C7EFF78F87966; miid=258847553226394565; thw=cn; UM_distinctid=15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578; _po=50014912_20.0_7_0.0_%E4%B8%8A%E6%B5%B7_%E4%B8%8A%E6%B5%B7_1_%E6%AC%BE%E5%88%B0%E5%8F%91%E8%B4%A7_10000_0_0_5_0_0_0.0_0.0_0.0_true_8940774860_null_0.0_20.0_null_null_null_135554976; x=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0; whl=-1%260%260%261490249632922; v=0; _tb_token_=fb3e5e7777336; uc3=sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWFmrrkZxc%2FsmA%3D&lg2=URm48syIIVrSKA%3D%3D; existShop=MTQ5MDI2MTk5OA%3D%3D; uss=BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D; lgc=qhn614; tracknick=qhn614; cookie2=41ad5598d9dd68684627ac5c611b83ac; sg=464; mt=np=&ci=21_1&cyk=-1_-1; cookie1=AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D; unb=135554976; skt=5540152fe95e625a; t=869b559af8caff3016c5e1c926316a66; _cc_=UIHiLt3xSw%3D%3D; tg=0; _l_g_=Ug%3D%3D; _nk_=qhn614; cookie17=UoLZWZQISEI7; cna=DpmyEIgHin4CAbaKZiKZKYni; l=Am1tPciTPShJuMJYvZjBZ3hJ/QPnkaGY; isg=Aqys--9e2KyZZsx6z0lIkLxafYpg41APZfHGaQbt5Nf-EU0bLnSZn2bjR25T; uc1=cookie16=VFC%2FuZ9az08KUQ56dCrZDlbNdA%3D%3D&cookie21=W5iHLLyFfXVRCJf5lG0u7A%3D%3D&cookie15=V32FPkk%2Fw0dUvg%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKK5imr3dA%3D%3D&tag=3&lng=zh_CN"
        }



        postdata = urllib.urlencode(values2)
        response = urllib.urlopen(u"https://upload.taobao.com/auction/publish/publish.htm", postdata)
        self.the_page = response.read()
        f=open("leimutest.html", "w")
        f.write(self.the_page)
        print (self.the_page)
        
        
    def send_multipart_post(self):
        newCookie={}
        for item in cookies:
          newCookie[item["name"]]=item["value"]
        body = {'mimeType':(None,'multipart/form-data; boundary=----WebKitFormBoundaryOtK5kELGehgIvjBB'),
                'text':('imagefile', open("./imagefile.xlsx",'r') ,'application/octet-stream')}

#        files = {'file': ('imagefile', open('./imagefile.xlsx', 'r'), 'application/octet-stream')}
        files = {'file': ('imagefile', open('./imagefile.xlsx', 'r'), 'multipart/form-data')}
        
        
        #, {'Expires': '0'})}

#        opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(self.use_cookie_file()))
#        opener = urllib2.build_opener()
        tuples = [(item["name"], item["value"]) for item in headers]
#        print(type(tuples))
#        print(tuples)
        dicts = {}
        for item in headers:
            dicts[item["name"]] = item["value"]

#        opener.addheaders = headers
#        req = urllib2.Request(upload_url, files)
#        response = opener.open(req)
#        print response.text
        s = requests.Session()
        dicts["Cookie"] = str(self.use_cookie_file())
        r = s.post(upload_url, files=body, headers=dicts, cookies=self.use_cookie_file())
#        r = s.post(upload_url, files=files, cookies=self.use_cookie_file())
#        print r.text
        

        
#            print((item["name"],item["value"]))
            
#        r=opener.open(login_url, body)    
#        r=opener.open(login_url, files)    
 #       print(r.read())
        
    def uploadfullwithcookies(self, title, cont, count):    
        values2 = {"Host": " upload.taobao.com",
            "Connection": " keep-alive",
            "Content-Length": " 14090",
            "Cache-Control": " max-age=0",
            "Origin": " https://upload.taobao.com",
            "Upgrade-Insecure-Requests": " 1",
            "User-Agent": " Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36",
            "Content-Type": " multipart/form-data; boundary=----WebKitFormBoundaryOtK5kELGehgIvjBB",
            "Accept": " text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
            "Referer": " https://upload.taobao.com/auction/publish/publish.htm",
            "Accept-Encoding": " gzip, deflate, br",
            "Accept-Language": " zh-CN,zh;q=0.8",
            "cookies":{
            "gm_item_upload_start_time":"1490261852625",
            "_umdata":"70CF403AFFD707DFED08E6D0D10782399C20737F7A4BC46D5700275F728DC4B0AD05CBF6A1A94BFBCD43AD3E795C914CC2C619865BC550D5090C7EFF78F87966",
            "gm_item_upload_take_time":"2417817",
            "miid":"258847553226394565",
            "thw":"cn",
            "UM_distinctid":"15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578",
            "_po":"50014912_20.0_7_0.0_%E4%B8%8A%E6%B5%B7_%E4%B8%8A%E6%B5%B7_1_%E6%AC%BE%E5%88%B0%E5%8F%91%E8%B4%A7_10000_0_0_5_0_0_0.0_0.0_0.0_true_8940774860_null_0.0_20.0_null_null_null_135554976",
            "x":"e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0",
            "whl":"-1%260%260%261490249632922",
            "v":"0",
            "_tb_token_":"fb3e5e7777336",
            "uc3":"sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWFmrrkZxc%2FsmA%3D&lg2=URm48syIIVrSKA%3D%3D",
            "existShop":"MTQ5MDI2MTk5OA%3D%3D",
            "uss":"BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D",
            "lgc":"qhn614",
            "tracknick":"qhn614",
            "cookie2":"41ad5598d9dd68684627ac5c611b83ac",
            "sg":"464",
            "mt":"np=&ci=21_1&cyk=-1_-1",
            "cookie1":"AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D",
            "unb":"135554976",
            "skt":"5540152fe95e625a",
            "t":"869b559af8caff3016c5e1c926316a66",
            "_cc_":"UIHiLt3xSw%3D%3D",
            "tg":"0",
            "_l_g_":"Ug%3D%3D",
            "_nk_":"qhn614",
            "cookie17":"UoLZWZQISEI7",
            "cna":"DpmyEIgHin4CAbaKZiKZKYni",
            "cookie16":"Vq8l%2BKCLySLZMFWHxqs8fwqnEw%3D%3D&cookie21=UtASsssmfaCOMId3WwGQmg%3D%3D&cookie15=UtASsssmOIJ0bQ%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKK5jXmvPg%3D%3D&tag=3&lng=zh_CN",
            "l":"AiYmhrZ6tlW-rXnt4oWaavTD9paprWrZ",
            "isg":"AmZmzec7AiIMO9Zs-euiLoIYt9xWwaoBS3N831AO-Al80wDtt9f6EUzhXXgl",
            }
        }
    
        values={
            "method": "POST",
            "accept-charset":"GBK",
            "url": "https://upload.taobao.com/auction/publish/publish.htm",
            "httpVersion": "HTTP/1.1",
            "headers": [
              {
                "name": "Cookie",
                "value": "_umdata=70CF403AFFD707DFED08E6D0D10782399C20737F7A4BC46D5700275F728DC4B0AD05CBF6A1A94BFBCD43AD3E795C914CC2C619865BC550D5090C7EFF78F87966; gm_item_upload_take_time=43919; miid=258847553226394565; thw=cn; UM_distinctid=15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578; x=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0; v=0; _tb_token_=fb3e5e7777336; uc3=sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWFnfbSjpUdxgM%3D&lg2=U%2BGCWk%2F75gdr5Q%3D%3D; existShop=MTQ5MDIzNTM0NQ%3D%3D; uss=BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D; lgc=qhn614; tracknick=qhn614; cookie2=41ad5598d9dd68684627ac5c611b83ac; sg=464; mt=np=&ci=21_1&cyk=-1_-1; cookie1=AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D; unb=135554976; skt=6085b89594eeb9e4; t=869b559af8caff3016c5e1c926316a66; _cc_=VFC%2FuZ9ajQ%3D%3D; tg=0; _l_g_=Ug%3D%3D; _nk_=qhn614; cookie17=UoLZWZQISEI7; cna=DpmyEIgHin4CAbaKZiKZKYni; _po=50014912_20.0_7_0.0_%E4%B8%8A%E6%B5%B7_%E4%B8%8A%E6%B5%B7_1_%E6%AC%BE%E5%88%B0%E5%8F%91%E8%B4%A7_10000_0_0_5_0_0_0.0_0.0_0.0_true_8940774860_null_0.0_20.0_null_null_null_135554976; l=As7OnrVH7t0GlcHlij2yko7Dnq6Qs5Ja; isg=AvT0IzCtEPQ46YRSR9EQOKSyxbLYSxi3bQkOQY5VfH8R-ZdDsNyYRxr7D4bb; uc1=cookie16=W5iHLLyFPlMGbLDwA%2BdvAGZqLg%3D%3D&cookie21=UIHiLt3xSixwH1aenGUFEQ%3D%3D&cookie15=U%2BGCWk%2F75gdr5Q%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKK8OXsyJA%3D%3D&tag=3&lng=zh_CN"
              },
              {
                "name": "Origin",
                "value": "https://upload.taobao.com"
              },
              {
                "name": "Accept-Encoding",
                "value": "gzip, deflate, br"
              },
              {
                "name": "Host",
                "value": "upload.taobao.com"
              },
              {
                "name": "Accept-Language",
                "value": "zh-CN,zh;q=0.8"
              },
              {
                "name": "Upgrade-Insecure-Requests",
                "value": "1"
              },
              {
                "name": "User-Agent",
                "value": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
              },
              {
                "name": "Content-Type",
                "value": "multipart/form-data; boundary=----WebKitFormBoundaryviR9IClzb1TjzhWV"
              },
              {
                "name": "Accept",
                "value": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8"
              },
              {
                "name": "Cache-Control",
                "value": "max-age=0"
              },
              {
                "name": "Referer",
                "value": "https://upload.taobao.com/auction/publish/publish.htm"
              },
              {
                "name": "Connection",
                "value": "keep-alive"
              },
              {
                "name": "Content-Length",
                "value": "14200"
              }
            ],
            "queryString": [],
            "cookies": [
              {
                "name": "_umdata",
                "value": "70CF403AFFD707DFED08E6D0D10782399C20737F7A4BC46D5700275F728DC4B0AD05CBF6A1A94BFBCD43AD3E795C914CC2C619865BC550D5090C7EFF78F87966",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "gm_item_upload_take_time",
                "value": "43919",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "miid",
                "value": "258847553226394565",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "thw",
                "value": "cn",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "UM_distinctid",
                "value": "15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "x",
                "value": "e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "v",
                "value": "0",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "_tb_token_",
                "value": "fb3e5e7777336",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "uc3",
                "value": "sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=Ev1FqNMR&id2=UoLZWZQISEI7&vt3=F8dARVWFnfbSjpUdxgM%3D&lg2=U%2BGCWk%2F75gdr5Q%3D%3D",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "existShop",
                "value": "MTQ5MDIzNTM0NQ%3D%3D",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "uss",
                "value": "BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "lgc",
                "value": "qhn614",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "tracknick",
                "value": "qhn614",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "cookie2",
                "value": "41ad5598d9dd68684627ac5c611b83ac",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "sg",
                "value": "464",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "mt",
                "value": "np=&ci=21_1&cyk=-1_-1",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "cookie1",
                "value": "AV10tdNqXVCfzGSiQssmwQqle8KnUSiB7qfbzOGrmNI%3D",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "unb",
                "value": "135554976",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "skt",
                "value": "6085b89594eeb9e4",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "t",
                "value": "869b559af8caff3016c5e1c926316a66",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "_cc_",
                "value": "VFC%2FuZ9ajQ%3D%3D",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "tg",
                "value": "0",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "_l_g_",
                "value": "Ug%3D%3D",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "_nk_",
                "value": "qhn614",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "cookie17",
                "value": "UoLZWZQISEI7",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "cna",
                "value": "DpmyEIgHin4CAbaKZiKZKYni",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "_po",
                "value": "50014912_20.0_7_0.0_%E4%B8%8A%E6%B5%B7_%E4%B8%8A%E6%B5%B7_1_%E6%AC%BE%E5%88%B0%E5%8F%91%E8%B4%A7_10000_0_0_5_0_0_0.0_0.0_0.0_true_8940774860_null_0.0_20.0_null_null_null_135554976",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "l",
                "value": "As7OnrVH7t0GlcHlij2yko7Dnq6Qs5Ja",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "isg",
                "value": "AvT0IzCtEPQ46YRSR9EQOKSyxbLYSxi3bQkOQY5VfH8R-ZdDsNyYRxr7D4bb",
                "expires": None,
                "httpOnly": False,
                "secure": False
              },
              {
                "name": "uc1",
                "value": "cookie16=W5iHLLyFPlMGbLDwA%2BdvAGZqLg%3D%3D&cookie21=UIHiLt3xSixwH1aenGUFEQ%3D%3D&cookie15=U%2BGCWk%2F75gdr5Q%3D%3D&existShop=true&pas=0&cookie14=UoW%2BuKK8OXsyJA%3D%3D&tag=3&lng=zh_CN",
                "expires": None,
                "httpOnly": False,
                "secure": False
              }
            ],
#            "headersSize": 2100,
#            "bodySize": 14200,
            "postData": {
              "mimeType": "multipart/form-data; boundary=----WebKitFormBoundaryviR9IClzb1TjzhWV",
              "text": "------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_tb_token_\"\r\n\r\nfb3e5e7777336\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"action\"\r\n\r\nupload/uploadAction\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isMImageUser\"\r\n\r\ntrue\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"sellerActionBeginTime\"\r\n\r\n1490236031224\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"publishPageCostTime\"\r\n\r\n1490236052004\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"browserTime\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"J_BySearch\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"J_SearchKey\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"J_ByLateCategory\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"J_ByCategorySystem\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ca\"\r\n\r\n50014912\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.i\"\r\n\r\nYmQ0ZTFkNGE4NDI3YTFiMjc2ZTgwY2MxZGNkYjdjYWUgYjkwODVjMDNkYTI4YWI5NjQyNjg1NmNiMzBlMGMyOTQgMTQ5MDIzNjAzMTIyNQ==\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.auc\"\r\n\r\n547148160013\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"itemNumId\"\r\n\r\n547148160013\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.cat\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.x\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ed\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isEdit\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"noticeParam\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"oldCat\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.fi\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.au\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.fr\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sho\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isn\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isol\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isg\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ish\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"catName\"\r\n\r\n½Å±¾/¹¤¾ß\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"selPPay\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"certType\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isZFBCertificatedUser\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"mainCategoryId\"\r\n\r\n50014811\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sp\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isb\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.a\"\r\n\r\nb\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.pa\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.b\"\r\n\r\n20\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.secure\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.iss\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isCodFlag\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"ccSn\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"ccIssuerDn\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"ccSignedData\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"userInnerShopId\"\r\n\r\n122254584\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"userOuterShopId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"userInnerShopSiteId\"\r\n\r\n4\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"userOuterShopSiteId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"wirelessmpChoosed\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isSellingLightningConsignment\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"lightningConsignment2\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"spuStoreId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ta\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"newProductTagId\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isFirst\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"tfsname\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.rel\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isLocalityLifeEtcItem\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isCustoemdSkuForbidDpcAuction\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"customCatId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"customCatName\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"distributionPageString\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isReselectCat\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isNeedCheckPictureCat\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"paymentSeting\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"forceUpdateWhenFakeCredit\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"wapDescEnable\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.stu\"\r\n\r\n5\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"auctionTypeInfoPrint\"\r\n\r\neditaVlue=false;canPublishNew=true;isOnlySecond=false;isRestAuction=;isBookStartLimit=false;isAcousticsLimit=false;isPrepayLimitCatAndSeller=false;is3GCatCanNewForSeller=;canPublishAlcoholNew=;noAgreement=;noPrepay=;shopSeller=true;isSellerShopeEnable=true;isShopRelease=;\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.auctio\"\r\n\r\n1048330278\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"customizedProp\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ti\"\r\n\r\n2017%s\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.pay\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.payv\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.m\"\r\n\r\n20\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.q\"\r\n\r\n10000\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.q\"\r\n\r\n10000\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"is_global_stock\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"gs_country\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"gs_type\"\r\n\r\n2\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ou\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"multimagefile\"; filename=\"\"\r\nContent-Type: application/octet-stream\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"videoAsPicThum\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"videoAsPicId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"videoAsPicDuration\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"pisAsVideoStatus\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl1\"\r\n\r\nhttps://img.alicdn.com/imgextra/i2/135554976/TB2l66KkohnpuFjSZFpXXcpuXXa_!!135554976.jpg\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n2\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl2\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n3\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl3\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n4\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl4\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n5\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl5\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.is\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.it\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.v\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sw\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"handleAuctionModule\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"modularizedArea\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.pc\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.d\"\r\n\r\n<p>%s</p>\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"state\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.w\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.wi\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.pc\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.w\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"pcTspeditorCid\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"itemNumId\"\r\n\r\n547148160013\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sh\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"shenbiPcDescHashCode\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"pcDescbackupEnable\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"wlDescbackupEnable\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isShowShenbiEditor\"\r\n\r\ntrue\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"TspWlsDesc\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.aut\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.auto\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.autoc\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.po\"\r\n\r\n8940774860\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ha\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.hav\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"new_prepay_tag\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isop\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.du\"\r\n\r\n7\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sta\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.auct\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_now\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"event_submit_do_publish\"\r\n\r\nanything\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"stamp\"\r\n\r\n1490235912297\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"type\"\r\n\r\nsubmit\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"submitType\"\r\n\r\najax\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV--\r\n" % (cont,title)
            }
        }
        
        pd = {
              "mimeType": "multipart/form-data; boundary=----WebKitFormBoundaryviR9IClzb1TjzhWV",
              "text": "------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_tb_token_\"\r\n\r\nfb3e5e7777336\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"action\"\r\n\r\nupload/uploadAction\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isMImageUser\"\r\n\r\ntrue\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"sellerActionBeginTime\"\r\n\r\n1490236031224\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"publishPageCostTime\"\r\n\r\n1490236052004\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"browserTime\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"J_BySearch\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"J_SearchKey\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"J_ByLateCategory\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"J_ByCategorySystem\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ca\"\r\n\r\n50014912\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.i\"\r\n\r\nYmQ0ZTFkNGE4NDI3YTFiMjc2ZTgwY2MxZGNkYjdjYWUgYjkwODVjMDNkYTI4YWI5NjQyNjg1NmNiMzBlMGMyOTQgMTQ5MDIzNjAzMTIyNQ==\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.auc\"\r\n\r\n547148160013\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"itemNumId\"\r\n\r\n547148160013\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.cat\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.x\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ed\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isEdit\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"noticeParam\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"oldCat\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.fi\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.au\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.fr\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sho\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isn\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isol\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isg\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ish\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"catName\"\r\n\r\n½Å±¾/¹¤¾ß\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"selPPay\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"certType\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isZFBCertificatedUser\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"mainCategoryId\"\r\n\r\n50014811\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sp\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isb\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.a\"\r\n\r\nb\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.pa\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.b\"\r\n\r\n20\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.secure\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.iss\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isCodFlag\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"ccSn\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"ccIssuerDn\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"ccSignedData\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"userInnerShopId\"\r\n\r\n122254584\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"userOuterShopId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"userInnerShopSiteId\"\r\n\r\n4\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"userOuterShopSiteId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"wirelessmpChoosed\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isSellingLightningConsignment\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"lightningConsignment2\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"spuStoreId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ta\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"newProductTagId\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isFirst\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"tfsname\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.rel\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isLocalityLifeEtcItem\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isCustoemdSkuForbidDpcAuction\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"customCatId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"customCatName\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"distributionPageString\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isReselectCat\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isNeedCheckPictureCat\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"paymentSeting\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"forceUpdateWhenFakeCredit\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"wapDescEnable\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.stu\"\r\n\r\n5\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"auctionTypeInfoPrint\"\r\n\r\neditaVlue=false;canPublishNew=true;isOnlySecond=false;isRestAuction=;isBookStartLimit=false;isAcousticsLimit=false;isPrepayLimitCatAndSeller=false;is3GCatCanNewForSeller=;canPublishAlcoholNew=;noAgreement=;noPrepay=;shopSeller=true;isSellerShopeEnable=true;isShopRelease=;\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.auctio\"\r\n\r\n1048330278\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"customizedProp\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ti\"\r\n\r\n2017%s\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.pay\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.payv\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.m\"\r\n\r\n20\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.q\"\r\n\r\n10000\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.q\"\r\n\r\n10000\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"is_global_stock\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"gs_country\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"gs_type\"\r\n\r\n2\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ou\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"multimagefile\"; filename=\"\"\r\nContent-Type: application/octet-stream\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"videoAsPicThum\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"videoAsPicId\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"videoAsPicDuration\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"pisAsVideoStatus\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n1\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl1\"\r\n\r\nhttps://img.alicdn.com/imgextra/i2/135554976/TB2l66KkohnpuFjSZFpXXcpuXXa_!!135554976.jpg\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n2\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl2\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n3\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl3\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n4\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl4\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"image_pos\"\r\n\r\n5\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"picUrl5\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.is\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.it\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.v\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sw\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"handleAuctionModule\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"modularizedArea\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.pc\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.d\"\r\n\r\n<p>%s</p>\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"state\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.w\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.wi\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.pc\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.w\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"pcTspeditorCid\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"itemNumId\"\r\n\r\n547148160013\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sh\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"shenbiPcDescHashCode\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"pcDescbackupEnable\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"wlDescbackupEnable\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"isShowShenbiEditor\"\r\n\r\ntrue\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"TspWlsDesc\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.aut\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.auto\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.autoc\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.po\"\r\n\r\n8940774860\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.ha\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.hav\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"new_prepay_tag\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.isop\"\r\n\r\nfalse\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.du\"\r\n\r\n7\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.sta\"\r\n\r\n\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_fma.pu._0.auct\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"_now\"\r\n\r\n0\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"event_submit_do_publish\"\r\n\r\nanything\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"stamp\"\r\n\r\n1490235912297\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"type\"\r\n\r\nsubmit\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV\r\nContent-Disposition: form-data; name=\"submitType\"\r\n\r\najax\r\n------WebKitFormBoundaryviR9IClzb1TjzhWV--\r\n" % (cont,title)
        }
        
        postdata = urllib.urlencode(pd)
        print(title)
        print(cont)
#        postdata = urllib.urlencode(pd)
#        postdata = urllib.urlencode(values)
        response = urllib.urlopen(u"https://upload.taobao.com/auction/publish/publish.htm", postdata)
        self.the_page = response.read()
        f=open("%d.html"%count,"w")
        f.write(self.the_page)
        print (self.the_page)
        
    def uploadquery(self):    
        values={
          "method": "GET",
          "url": "https://gm.mmstat.com/mjzx.50.1?catid=50014912&status=1&timespans=191846&cache=1490176912888",
          "httpVersion": "unknown",
          "headers": [
            {
              "name": "Referer",
              "value": "https://upload.taobao.com/auction/publish/publish.htm"
            },
            {
              "name": "User-Agent",
              "value": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
            }
          ],
          "queryString": [
            {
              "name": "catid",
              "value": "50014912"
            },
            {
              "name": "status",
              "value": "1"
            },
            {
              "name": "timespans",
              "value": "191846"
            },
            {
              "name": "cache",
              "value": "1490176912888"
            }
          ],
          "cookies": [],
          "headersSize": -1,
          "bodySize": 0
        }
        postdata = urllib.urlencode(values)
        response = urllib.urlopen(baobei_leimu, postdata)
        self.the_page = response.read()
        print (self.the_page)
    
    def read_baobei(self, str):
        dict = {}
        f = open(str, "r")
        lines = f.readlines()
        count = 0
        for line in lines:
            if (line!="" or line!=null):
                print(line)
                list = line.split("qqq")
                if len(list)==1:
                    continue
                count=count+1
                for item in list:
                    print(item)
                    dict["content%d"%count]=list[0]
                    dict["title%d"%count]=list[1]
        return dict

    def click_test(self):
        values = {"id":"547005033344&amp","ns":"1&amp","abbucket":"19#detail"}
        postData = urllib.urlencode(values)
        response = urllib.urlopen("https//item.taobao.com/item.htm?%s"%postData)
        self.the_page = response.read()
        print (self.the_page)

        
    def click_baobei(self):
        values = {
          "method": "GET",
          "httpVersion": "HTTP/1.1",
          "headers": [
            {
              "name": "Cookie",
              "value": "miid=258847553226394565; thw=cn; UM_distinctid=15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578; uss=BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D; cookie2=41ad5598d9dd68684627ac5c611b83ac; t=869b559af8caff3016c5e1c926316a66; _cc_=VFC%2FuZ9ajQ%3D%3D; tg=0; x=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0; whl=-1%260%260%261490249632922; uc3=sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=&id2=&lg2=; tracknick=; mt=ci=0_0&cyk=-1_-1; v=0; alitrackid=www.taobao.com; lastalitrackid=www.taobao.com; _tb_token_=fb3e5e7777336; cna=DpmyEIgHin4CAbaKZiKZKYni; uc1=cookie14=UoW%2BuKK6tzgG4g%3D%3D; JSESSIONID=1BB252467D70FCA3E2AF22E63CF1448C; l=AqengQekR1oXJljKy/5rNSVMt9Fzp3sU; isg=AiIikZnVjq5gM5KY5b9e0mb0c6hHsROq74-4C2y5xRVaP8S5VgKknLz5Gcw5"
            },
            {
              "name": "Accept-Encoding",
              "value": "gzip, deflate, sdch, br"
            },
            {
              "name": "Host",
              "value": "s.taobao.com"
            },
            {
              "name": "Accept-Language",
              "value": "zh-CN,zh;q=0.8"
            },
            {
              "name": "User-Agent",
              "value": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
            },
            {
              "name": "Accept",
              "value": "text/javascript, application/javascript, application/ecmascript, application/x-ecmascript, */*; q=0.01"
            },
            {
              "name": "Referer",
              "value": "https://s.taobao.com/search?q=%E4%BB%A3%E6%89%BE%E7%94%B5%E5%AD%90%E4%B9%A6&imgfile=&js=1&stats_click=search_radio_all%3A1&initiative_id=staobaoz_20170323&ie=utf8&bcoffset=4&ntoffset=4&p4ppushleft=1%2C48&loc=%E4%B8%8A%E6%B5%B7&filter=reserve_price%5B%2C10%5D"
            },
            {
              "name": "X-Requested-With",
              "value": "XMLHttpRequest"
            },
            {
              "name": "Connection",
              "value": "keep-alive"
            }
          ],
          "queryString": [
            {
              "name": "_ksTS",
              "value": "1490249940413_1300"
            },
            {
              "name": "callback",
              "value": "jsonp1301"
            },
            {
              "name": "ajax",
              "value": "true"
            },
            {
              "name": "app",
              "value": "api"
            },
            {
              "name": "m",
              "value": "postcustomized"
            },
            {
              "name": "q",
              "value": "%E4%BB%A3%E6%89%BE%E7%94%B5%E5%AD%90%E4%B9%A6"
            },
            {
              "name": "s",
              "value": "null"
            },
            {
              "name": "nid",
              "value": "547238847197"
            }
          ],
          "cookies": [
            {
              "name": "miid",
              "value": "258847553226394565",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "thw",
              "value": "cn",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "UM_distinctid",
              "value": "15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uss",
              "value": "BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cookie2",
              "value": "41ad5598d9dd68684627ac5c611b83ac",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "t",
              "value": "869b559af8caff3016c5e1c926316a66",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_cc_",
              "value": "VFC%2FuZ9ajQ%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "tg",
              "value": "0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "x",
              "value": "e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "whl",
              "value": "-1%260%260%261490249632922",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uc3",
              "value": "sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=&id2=&lg2=",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "tracknick",
              "value": "",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "mt",
              "value": "ci=0_0&cyk=-1_-1",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "v",
              "value": "0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "alitrackid",
              "value": "www.taobao.com",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "lastalitrackid",
              "value": "www.taobao.com",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_tb_token_",
              "value": "fb3e5e7777336",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cna",
              "value": "DpmyEIgHin4CAbaKZiKZKYni",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uc1",
              "value": "cookie14=UoW%2BuKK6tzgG4g%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "JSESSIONID",
              "value": "1BB252467D70FCA3E2AF22E63CF1448C",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "l",
              "value": "AqengQekR1oXJljKy/5rNSVMt9Fzp3sU",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "isg",
              "value": "AiIikZnVjq5gM5KY5b9e0mb0c6hHsROq74-4C2y5xRVaP8S5VgKknLz5Gcw5",
              "expires": None,
              "httpOnly": False,
              "secure": False
            }
          ],
          "headersSize": 1644,
          "bodySize": 0
        }
        postdata = urllib.urlencode(values)
        response = urllib.urlopen(u"https://s.taobao.com/api?%s" % postdata)
        self.the_page = response.read()
        print (self.the_page)
        
        '''values = {
          "method": "GET",
          "httpVersion": "HTTP/1.1",
          "headers": [
            {
              "name": "Cookie",
              "value": "miid=258847553226394565; thw=cn; UM_distinctid=15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578; uss=BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D; cookie2=41ad5598d9dd68684627ac5c611b83ac; t=869b559af8caff3016c5e1c926316a66; _cc_=VFC%2FuZ9ajQ%3D%3D; tg=0; x=e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0; whl=-1%260%260%261490249632922; uc3=sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=&id2=&lg2=; tracknick=; mt=ci=0_0&cyk=-1_-1; v=0; alitrackid=www.taobao.com; lastalitrackid=www.taobao.com; _tb_token_=fb3e5e7777336; cna=DpmyEIgHin4CAbaKZiKZKYni; uc1=cookie14=UoW%2BuKK6tzgG4g%3D%3D; JSESSIONID=1BB252467D70FCA3E2AF22E63CF1448C; l=AqengQekR1oXJljKy/5rNSVMt9Fzp3sU; isg=AiIikZnVjq5gM5KY5b9e0mb0c6hHsROq74-4C2y5xRVaP8S5VgKknLz5Gcw5"
            },
            {
              "name": "Accept-Encoding",
              "value": "gzip, deflate, sdch, br"
            },
            {
              "name": "Host",
              "value": "s.taobao.com"
            },
            {
              "name": "Accept-Language",
              "value": "zh-CN,zh;q=0.8"
            },
            {
              "name": "User-Agent",
              "value": "Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36"
            },
            {
              "name": "Accept",
              "value": "text/javascript, application/javascript, application/ecmascript, application/x-ecmascript, */*; q=0.01"
            },
            {
              "name": "Referer",
              "value": "https://s.taobao.com/search?q=%E4%BB%A3%E6%89%BE%E7%94%B5%E5%AD%90%E4%B9%A6&imgfile=&js=1&stats_click=search_radio_all%3A1&initiative_id=staobaoz_20170323&ie=utf8&bcoffset=4&ntoffset=4&p4ppushleft=1%2C48&loc=%E4%B8%8A%E6%B5%B7&filter=reserve_price%5B%2C10%5D"
            },
            {
              "name": "X-Requested-With",
              "value": "XMLHttpRequest"
            },
            {
              "name": "Connection",
              "value": "keep-alive"
            }
          ],
          "queryString": [
            {
              "name": "_ksTS",
              "value": "1490249940413_1300"
            },
            {
              "name": "callback",
              "value": "jsonp1301"
            },
            {
              "name": "ajax",
              "value": "true"
            },
            {
              "name": "app",
              "value": "api"
            },
            {
              "name": "m",
              "value": "postcustomized"
            },
            {
              "name": "q",
              "value": "%E4%BB%A3%E6%89%BE%E7%94%B5%E5%AD%90%E4%B9%A6"
            },
            {
              "name": "s",
              "value": "null"
            },
            {
              "name": "nid",
              "value": "547238847197"
            }
          ],
          "cookies": [
            {
              "name": "miid",
              "value": "258847553226394565",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "thw",
              "value": "cn",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "UM_distinctid",
              "value": "15ac57fe04456c-0caf936e65f7fe-6a11157a-a41c3-15ac57fe045578",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uss",
              "value": "BYXLXfubmhhu%2BLZdsrSYQbT1BAFhcgaem3eMdpLBN8MSiGMotZA79MMl%2B9s%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cookie2",
              "value": "41ad5598d9dd68684627ac5c611b83ac",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "t",
              "value": "869b559af8caff3016c5e1c926316a66",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_cc_",
              "value": "VFC%2FuZ9ajQ%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "tg",
              "value": "0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "x",
              "value": "e%3D1%26p%3D*%26s%3D0%26c%3D0%26f%3D0%26g%3D0%26t%3D0%26__ll%3D-1%26_ato%3D0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "whl",
              "value": "-1%260%260%261490249632922",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uc3",
              "value": "sg2=AQcgxLTieJjRk4V3dcCwBsHtwcnIFiu7tKcbLSxIJdg%3D&nk2=&id2=&lg2=",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "tracknick",
              "value": "",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "mt",
              "value": "ci=0_0&cyk=-1_-1",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "v",
              "value": "0",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "alitrackid",
              "value": "www.taobao.com",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "lastalitrackid",
              "value": "www.taobao.com",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "_tb_token_",
              "value": "fb3e5e7777336",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "cna",
              "value": "DpmyEIgHin4CAbaKZiKZKYni",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "uc1",
              "value": "cookie14=UoW%2BuKK6tzgG4g%3D%3D",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "JSESSIONID",
              "value": "1BB252467D70FCA3E2AF22E63CF1448C",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "l",
              "value": "AqengQekR1oXJljKy/5rNSVMt9Fzp3sU",
              "expires": None,
              "httpOnly": False,
              "secure": False
            },
            {
              "name": "isg",
              "value": "AiIikZnVjq5gM5KY5b9e0mb0c6hHsROq74-4C2y5xRVaP8S5VgKknLz5Gcw5",
              "expires": None,
              "httpOnly": False,
              "secure": False
            }
          ],
          "headersSize": 1644,
          "bodySize": 0
        }
        postdata = urllib.urlencode(values)
        response = urllib.urlopen("https://s.taobao.com/api?%s",postdata)
        self.the_page = response.read()
        print (self.the_page)'''

    def total_push(self):
        dicts = entry.read_baobei("./baobei100.txt")
        #time.sleep(10)
        size = len(dicts)
        size = size/2
        for i in range(1,size):
            entry.uploadfullwithcookies(dicts["title%d"%i], dicts["content%d"%i], i)
            print(dicts["title%d"%i])
            print(dicts["content%d"%i])
            time.sleep(60)

    def get_cookie_withurllib(self):
        filename = '/cookie2.txt'
        cookie_in = cookielib.MozillaCookieJar(filename)
        handler = urllib2.HTTPCookieProcessor(cookie_in)                        
        opener = urllib2.build_opener(handler)
        postdata = urllib.urlencode({'TPL_username':'qhn614','TPL_password':'taobao161230@@'})
        response = opener.open(login_url,postdata)
        print response.read()
        cookie_in.save(ignore_discard=True, ignore_expires=True)

    def use_cookie_file(self):
        cookie_out = cookielib.MozillaCookieJar()
        cookie_out.load('./cookie.txt', ignore_discard=True, ignore_expires=True)
        print(cookie_out)
        return cookie_out
        
    def writexls(self):
        dest_filename = 'imagefile.xlsx'        
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "imagefile"
        count = 0
        for item in post_data:
            count = count + 1
            ws1.cell(column=1, row=count,value=str(item["name"]))
            ws1.cell(column=2, row=count, value=str(item["value"]))
        wb.save(filename = dest_filename)    
        
    def writetxt(self):
        f = open("imagefile.txt",'w')
        f.write(str(post_data))
        
if __name__ == "__main__":
    entry = EntryDemo()
#    entry.writexls()
    entry.get_cookie_withurllib()
    
        
        
        
